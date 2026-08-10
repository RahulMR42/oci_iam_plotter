"""Tests for portable report export formats."""

from copy import deepcopy
from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader

from oci_iam_plotter.analysis import find_duplicates, policy_analysis
from oci_iam_plotter.reporting import csv_report, markdown_report, pdf_report, report_payload, write_report, xlsx_report


def test_csv_report_contains_inventory_and_user_access(snapshot) -> None:
    payload = report_payload(snapshot, policy_analysis(snapshot, "user-1"), find_duplicates(snapshot), None)
    content = csv_report(payload)
    assert "section,category,name,value,confidence" in content
    assert "inventory,entity_count,user,1,direct" in content
    assert "user_access,implied_permission,alice,read all-resources (tenancy),inferred" in content


def test_xlsx_report_has_audit_sheets_and_styled_headers(snapshot) -> None:
    payload = report_payload(snapshot, policy_analysis(snapshot, "user-1"), find_duplicates(snapshot), None)
    data = xlsx_report(snapshot, payload)
    workbook = load_workbook(BytesIO(data), read_only=False)
    assert workbook.sheetnames == ["Summary", "Entities", "Relationships", "Memberships", "Policy Statements", "User Access"]
    assert workbook["Summary"].freeze_panes == "A2"
    assert workbook["Entities"]["A1"].font.bold
    assert workbook["User Access"]["A2"].value == "user"


def test_write_report_supports_csv_and_xlsx(snapshot, tmp_path) -> None:
    payload = report_payload(snapshot, None, find_duplicates(snapshot), None)
    csv_path = write_report(payload, tmp_path / "report.csv", snapshot)
    xlsx_path = write_report(payload, tmp_path / "report.xlsx", snapshot)
    assert csv_path.read_text(encoding="utf-8").startswith("section,category")
    assert xlsx_path.read_bytes().startswith(b"PK")


def test_multi_user_markdown_and_pdf_are_tabular(snapshot) -> None:
    alice = policy_analysis(snapshot, "user-1")
    bob = deepcopy(alice)
    bob["user"] = {**alice["user"], "id": "user-2", "name": "bob"}
    bob["groups"] = []
    bob["applicable_policy_statements"] = []
    bob["implied_permissions"] = []
    payload = report_payload(snapshot, [alice, bob], find_duplicates(snapshot), None)

    markdown = markdown_report(payload)
    assert "| User | Groups | Matching policies |" in markdown
    assert "| alice | 1 | 1 | 1 |" in markdown
    assert "| bob | 0 | 0 | 0 |" in markdown

    pdf = pdf_report(payload)
    reader = PdfReader(BytesIO(pdf))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert pdf.startswith(b"%PDF")
    assert "User access comparison" in text
    assert "alice" in text and "bob" in text


def test_write_report_supports_pdf(snapshot, tmp_path) -> None:
    payload = report_payload(snapshot, policy_analysis(snapshot, "user-1"), find_duplicates(snapshot), None)
    pdf_path = write_report(payload, tmp_path / "report.pdf", snapshot)
    assert pdf_path.read_bytes().startswith(b"%PDF")
