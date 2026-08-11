"""Structured JSON and portable human-readable report generation."""

from __future__ import annotations

from dataclasses import asdict
import csv
from html import escape as html_escape
from io import BytesIO, StringIO
import json
from pathlib import Path
from typing import Any

from .models import Snapshot


def report_payload(
    snapshot: Snapshot,
    access: dict | list[dict] | None,
    duplicates: dict,
    summary: dict | list[dict] | None,
    risk_posture: dict | None = None,
) -> dict[str, Any]:
    """Assemble a machine-readable report payload from local analysis outputs."""
    counts: dict[str, int] = {}
    for entity in snapshot.entities:
        counts[entity.kind] = counts.get(entity.kind, 0) + 1
    relationship_counts: dict[str, int] = {}
    for relationship in snapshot.relationships:
        relationship_counts[relationship.kind] = relationship_counts.get(relationship.kind, 0) + 1
    access_analyses = access if isinstance(access, list) else ([access] if access else [])
    summaries = summary if isinstance(summary, list) else ([summary] if summary else [])
    attributed_summaries = []
    for index, item in enumerate(summaries):
        user = next((analysis.get("user", {}) for analysis in access_analyses
                     if analysis.get("user", {}).get("id") == item.get("user_id")), None)
        user = user or (access_analyses[index].get("user", {}) if index < len(access_analyses) else {})
        attributed_summaries.append({"user_id": user.get("id"), "user_name": user.get("name"), **item})
    return {"snapshot": {"tenancy_id": snapshot.tenancy_id, "collected_at": snapshot.collected_at,
                         "source_hash": snapshot.source_hash, "warnings": snapshot.warnings},
            "inventory": counts, "relationships": relationship_counts,
            # Keep singular keys for downstream consumers created before multi-user reports.
            "access_analysis": access_analyses[0] if len(access_analyses) == 1 else None,
            "access_analyses": access_analyses,
            "duplicates": duplicates,
            "summary": attributed_summaries[0] if len(attributed_summaries) == 1 else None,
            "summaries": attributed_summaries, "risk_posture": risk_posture or {}}


def _access_analyses(payload: dict[str, Any]) -> list[dict[str, Any]]:
    """Return normalized user analyses from new or legacy report payloads."""
    analyses = payload.get("access_analyses")
    if analyses is not None:
        return analyses
    legacy = payload.get("access_analysis")
    return [legacy] if legacy else []


def _markdown_cell(value: Any) -> str:
    """Escape a compact value for a Markdown table cell."""
    return str(value).replace("|", "\\|").replace("\n", " ")


def markdown_report(payload: dict[str, Any]) -> str:
    """Render an audit-oriented Markdown report without credentials or raw SDK data."""
    lines = ["# OCI IAM Plotter report", "", f"- Tenancy: `{payload['snapshot']['tenancy_id']}`", f"- Collected: {payload['snapshot']['collected_at']}", "", "## Entity inventory", ""]
    lines += [f"- {kind}: {count}" for kind, count in sorted(payload["inventory"].items())]
    risk_posture = payload.get("risk_posture", {})
    if risk_posture:
        lines += ["", "## Access risk posture", ""]
        lines += [f"- {level.title()}: {count}" for level, count in risk_posture.get("distribution", {}).items()]
        lines += ["", "| User | Risk | Score | Policy-based signals |", "|---|---|---:|---|"]
        for item in risk_posture.get("flagged_users", []):
            signals = "; ".join(signal["permission"] for signal in item.get("signals", [])[:3]) or "Review matched policy evidence"
            lines.append(f"| {_markdown_cell(item['name'])} | {item['risk_level'].title()} | {item['risk_score']} | {_markdown_cell(signals)} |")
        lines += ["", "### Top risk elements", "", "| User | Risk | Score | Permission | Why it was prioritized |", "|---|---|---:|---|---|"]
        for item in risk_posture.get("top_risk_elements", [])[:20]:
            lines.append("| " + " | ".join(_markdown_cell(value) for value in (
                item["user"], item["level"].title(), item["score"], item["permission"], "; ".join(item.get("reasons", [])))) + " |")
        lines += ["", risk_posture.get("method", "")]
    lines += ["", "## Relationship inventory", ""]
    lines += [f"- {kind}: {count}" for kind, count in sorted(payload.get("relationships", {}).items())] or ["- None"]
    analyses = _access_analyses(payload)
    if analyses:
        lines += ["", "## User access comparison", "",
                  "| User | Groups | Matching policies | Implied permissions | Relevant ambiguities | Confidence |",
                  "|---|---:|---:|---:|---:|---|"]
        for access in analyses:
            policies = {item.get("policy_id") for item in access.get("applicable_policy_statements", [])}
            lines.append("| " + " | ".join(_markdown_cell(item) for item in (
                access["user"]["name"], len(access.get("groups", [])), len(policies),
                len(access.get("implied_permissions", [])),
                len(access.get("unresolved_ambiguous_statements", [])), access.get("confidence", "inferred"))) + " |")
        lines += ["", "## Per-user access details"]
        summaries = {item.get("user_id"): item for item in payload.get("summaries", [])}
        for access in analyses:
            user = access["user"]
            lines += ["", f"### {_markdown_cell(user['name'])}", "",
                      "| Evidence | Details | Confidence |", "|---|---|---|",
                      f"| User | `{_markdown_cell(user['id'])}` | Direct |",
                      f"| Groups | {_markdown_cell(', '.join(group['name'] for group in access.get('groups', [])) or 'None')} | Direct |",
                      f"| Administrator role | {_markdown_cell(', '.join(access.get('administrator_roles', [])) or 'None')} | Group membership |",
                      f"| Implied permissions | {_markdown_cell('; '.join(access.get('implied_permissions', [])) or 'None matched')} | Inferred |",
                      f"| Relevant ambiguous statements | {len(access.get('unresolved_ambiguous_statements', []))} | Requires review |",
                      "", "#### Confidence and limitations", ""]
            lines += [f"- {item}" for item in access.get("limitations", [])] or ["- No additional limitations recorded."]
            if user.get("id") in summaries:
                lines += ["", "#### Generated summary", "", summaries[user["id"]]["summary"]]
    duplicates = payload["duplicates"]
    lines += ["", "## Duplicate / overlap candidates", "", f"- Exact entity-name candidates: {len(duplicates['exact_entity_name_candidates'])}", f"- Exact policy-statement candidates: {len(duplicates['exact_policy_statement_candidates'])}", f"- Near-name candidates: {len(duplicates['near_entity_name_candidates'])}", "", "## Next checks", "", "- Validate candidate access with OCI policy evaluation and service-specific controls.", "- Review conditional statements and policies outside the collected scope if needed."]
    return "\n".join(lines) + "\n"


def html_report(payload: dict[str, Any]) -> str:
    """Render the canonical Markdown report as a self-contained HTML document."""
    markdown = markdown_report(payload)
    lines, parts, index = markdown.splitlines(), [], 0
    def cells(line: str) -> list[str]:
        return [html_escape(cell.strip().replace("\\|", "|")) for cell in line.strip().strip("|").split("|")]
    while index < len(lines):
        line = lines[index]
        if not line.strip(): index += 1; continue
        if line.startswith("### "): parts.append(f"<h3>{html_escape(line[4:])}</h3>"); index += 1; continue
        if line.startswith("## "): parts.append(f"<h2>{html_escape(line[3:])}</h2>"); index += 1; continue
        if line.startswith("# "): parts.append(f"<h1>{html_escape(line[2:])}</h1>"); index += 1; continue
        if line.startswith("- "):
            values = []
            while index < len(lines) and lines[index].startswith("- "):
                values.append(f"<li>{html_escape(lines[index][2:])}</li>"); index += 1
            parts.append("<ul>" + "".join(values) + "</ul>"); continue
        if "|" in line and index + 1 < len(lines) and lines[index + 1].replace("|", "").replace("-", "").replace(":", "").strip() == "":
            headers = cells(line); index += 2; rows = []
            while index < len(lines) and "|" in lines[index] and lines[index].strip():
                rows.append("<tr>" + "".join(f"<td>{cell}</td>" for cell in cells(lines[index])) + "</tr>"); index += 1
            parts.append("<table><thead><tr>" + "".join(f"<th>{cell}</th>" for cell in headers) + "</tr></thead><tbody>" + "".join(rows) + "</tbody></table>"); continue
        paragraph = []
        while index < len(lines) and lines[index].strip() and not lines[index].startswith(("#", "- ")):
            paragraph.append(lines[index]); index += 1
        parts.append(f"<p>{html_escape(' '.join(paragraph))}</p>")
    style = "body{font:15px/1.55 Inter,Arial,sans-serif;color:#27221e;max-width:1100px;margin:32px auto;padding:0 24px}h1,h2,h3{color:#17365d}h2{border-bottom:1px solid #ddd;padding-bottom:6px;margin-top:32px}table{border-collapse:collapse;width:100%;margin:14px 0}th{background:#17365d;color:#fff}th,td{padding:8px;border:1px solid #cfc8c0;text-align:left;vertical-align:top}tr:nth-child(even){background:#f7f5f2}"
    return f"<!doctype html><html><head><meta charset=\"utf-8\"><title>OCI IAM Plotter report</title><style>{style}</style></head><body>{''.join(parts)}</body></html>"


def csv_report(payload: dict[str, Any]) -> str:
    """Render a flat, portable audit summary CSV."""
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["section", "category", "name", "value", "confidence"])
    for name, value in payload["snapshot"].items():
        writer.writerow(["snapshot", "metadata", name, json.dumps(value) if isinstance(value, (list, dict)) else value, "direct"])
    for name, value in sorted(payload.get("inventory", {}).items()):
        writer.writerow(["inventory", "entity_count", name, value, "direct"])
    for name, value in sorted(payload.get("relationships", {}).items()):
        writer.writerow(["relationships", "relationship_count", name, value, "direct_or_derived"])
    for access in _access_analyses(payload):
        writer.writerow(["user_access", "user", access["user"]["name"], access["user"]["id"], access.get("confidence")])
        for group in access.get("groups", []):
            writer.writerow(["user_access", "group", group["name"], group["id"], "direct"])
        for permission in access.get("implied_permissions", []):
            writer.writerow(["user_access", "implied_permission", access["user"]["name"], permission, "inferred"])
        risk = access.get("risk", {})
        writer.writerow(["user_access", "risk_level", access["user"]["name"], risk.get("level", "low"), "heuristic"])
    return output.getvalue()


def xlsx_report(snapshot: Snapshot, payload: dict[str, Any]) -> bytes:
    """Build a formatted multi-sheet Excel workbook for audit investigation."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(title: str, rows: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet(title)
        if not rows:
            sheet.append(["No records"])
            return
        headers = list(rows[0])
        sheet.append(headers)
        for row in rows:
            sheet.append([_excel_value(row.get(header)) for header in headers])
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="top")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            width = min(48, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
            for cell in column[1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary_rows = [
        {"section": "snapshot", "name": key, "value": value}
        for key, value in payload["snapshot"].items()
    ] + [
        {"section": "entity inventory", "name": key, "value": value}
        for key, value in sorted(payload.get("inventory", {}).items())
    ] + [
        {"section": "relationship inventory", "name": key, "value": value}
        for key, value in sorted(payload.get("relationships", {}).items())
    ]
    add_sheet("Summary", summary_rows)
    add_sheet("Entities", [asdict(item) for item in snapshot.entities])
    add_sheet("Relationships", [asdict(item) for item in snapshot.relationships])
    add_sheet("Memberships", [asdict(item) for item in snapshot.memberships])
    add_sheet("Policy Statements", [asdict(item) for item in snapshot.statements])
    analyses = _access_analyses(payload)
    if analyses:
        rows = []
        for access in analyses:
            rows.append({"record_type": "user", "name": access["user"]["name"], "value": access["user"]["id"],
                         "confidence": access.get("confidence")})
            rows += [{"record_type": "group", "name": item["name"], "value": item["id"], "confidence": "direct"}
                     for item in access.get("groups", [])]
            rows += [{"record_type": "implied_permission", "name": access["user"]["name"], "value": item,
                      "confidence": "inferred"} for item in access.get("implied_permissions", [])]
        add_sheet("User Access", rows)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _excel_value(value: Any) -> Any:
    """Convert nested normalized data to an Excel-safe scalar."""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, sort_keys=True, default=str)
    return value


def pdf_report(payload: dict[str, Any]) -> bytes:
    """Build a paginated audit PDF with a multi-user comparison and evidence tables."""
    from xml.sax.saxutils import escape

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import KeepTogether, LongTable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    stream = BytesIO()
    document = SimpleDocTemplate(
        stream, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm,
        topMargin=18 * mm, bottomMargin=17 * mm,
        title="OCI IAM Plotter report", author="OCI IAM Plotter",
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], textColor=colors.HexColor("#17365D"),
                              fontSize=20, leading=24, spaceAfter=8))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], textColor=colors.HexColor("#1F4E78"),
                              fontSize=13, leading=16, spaceBefore=10, spaceAfter=6))
    styles.add(ParagraphStyle(name="Subsection", parent=styles["Heading3"], textColor=colors.HexColor("#1F4E78"),
                              fontSize=11, leading=14, spaceBefore=8, spaceAfter=4))
    styles.add(ParagraphStyle(name="Cell", parent=styles["BodyText"], fontSize=7.5, leading=9.5))
    styles.add(ParagraphStyle(name="CellHeader", parent=styles["BodyText"], fontSize=7.5, leading=9,
                              textColor=colors.white, alignment=TA_CENTER))

    def paragraph(value: Any, style: str = "Cell") -> Paragraph:
        text = escape(str(value if value not in (None, "") else "-"))
        return Paragraph(text, styles[style])

    def table(rows: list[list[Any]], widths: list[float]) -> Table:
        rendered = [[paragraph(value, "CellHeader" if row_index == 0 else "Cell")
                     for value in row] for row_index, row in enumerate(rows)]
        result = LongTable(rendered, colWidths=widths, repeatRows=1, splitByRow=1)
        result.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F6F8FA")),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B8C2CC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return result

    story: list[Any] = [Paragraph("OCI IAM Plotter report", styles["ReportTitle"]),
                        Paragraph(f"Tenancy: {escape(str(payload['snapshot']['tenancy_id']))}", styles["BodyText"]),
                        Paragraph(f"Collected: {escape(str(payload['snapshot']['collected_at']))}", styles["BodyText"]),
                        Spacer(1, 4 * mm), Paragraph("Entity inventory", styles["Section"])]
    inventory_rows = [["Entity type", "Count"], *[[key.replace("_", " "), value]
                                                    for key, value in sorted(payload.get("inventory", {}).items())]]
    story += [table(inventory_rows, [125 * mm, 45 * mm]),
              Paragraph("Access risk posture", styles["Section"])]
    risk_posture = payload.get("risk_posture", {})
    if risk_posture:
        risk_rows = [["Risk level", "Users"], *[[level.title(), count] for level, count in risk_posture.get("distribution", {}).items()]]
        story += [table(risk_rows, [125 * mm, 45 * mm]),
                  Paragraph(str(risk_posture.get("method", "")), styles["BodyText"])]
        flagged_rows = [["User", "Risk", "Score", "Signals"]]
        for item in risk_posture.get("flagged_users", []):
            flagged_rows.append([item["name"], item["risk_level"].title(), item["risk_score"],
                                 "; ".join(signal["permission"] for signal in item.get("signals", [])[:3])])
        if len(flagged_rows) > 1:
            story += [Paragraph("Prioritized review", styles["Subsection"]), table(flagged_rows, [36 * mm, 24 * mm, 18 * mm, 92 * mm])]
        top_rows = [["User", "Risk", "Score", "Permission", "Why prioritized"]]
        for item in risk_posture.get("top_risk_elements", [])[:20]:
            top_rows.append([item["user"], item["level"].title(), item["score"], item["permission"],
                             "; ".join(item.get("reasons", []))])
        if len(top_rows) > 1:
            story += [Paragraph("Top risk elements", styles["Subsection"]),
                      table(top_rows, [28 * mm, 18 * mm, 14 * mm, 42 * mm, 68 * mm])]
    story.append(Paragraph("User access comparison", styles["Section"]))
    analyses = _access_analyses(payload)
    comparison = [["User", "Groups", "Policies", "Permissions", "Ambiguous", "Confidence"]]
    for access in analyses:
        policies = {item.get("policy_id") for item in access.get("applicable_policy_statements", [])}
        comparison.append([access["user"]["name"], len(access.get("groups", [])), len(policies),
                           len(access.get("implied_permissions", [])),
                           len(access.get("unresolved_ambiguous_statements", [])), access.get("confidence", "inferred")])
    if analyses:
        story.append(table(comparison, [42 * mm, 20 * mm, 22 * mm, 27 * mm, 27 * mm, 32 * mm]))
    else:
        story.append(Paragraph("No users were selected for access analysis.", styles["BodyText"]))

    summaries = {item.get("user_id"): item for item in payload.get("summaries", [])}
    for access in analyses:
        user = access["user"]
        story += [Paragraph(user["name"], styles["Subsection"])]
        details = [
            ["Evidence", "Details", "Confidence"],
            ["User", user["id"], "Direct"],
            ["Groups", ", ".join(item["name"] for item in access.get("groups", [])) or "None", "Direct"],
            ["Risk posture", f"{access.get('risk', {}).get('level', 'low').title()} ({access.get('risk', {}).get('score', 0)})", "Heuristic"],
            ["Implied permissions", "; ".join(access.get("implied_permissions", [])) or "None matched", "Inferred"],
            ["Relevant ambiguities", len(access.get("unresolved_ambiguous_statements", [])), "Requires review"],
        ]
        story += [table(details, [40 * mm, 95 * mm, 35 * mm])]
        limitations = access.get("limitations", [])
        if limitations:
            limitation_text = "; ".join(escape(str(item)) for item in limitations)
            story.append(Paragraph(f"<b>Limitations:</b> {limitation_text}", styles["Cell"]))
        generated = summaries.get(user.get("id"))
        if generated:
            story += [Paragraph("Generated summary", styles["Subsection"]),
                      Paragraph(escape(str(generated["summary"])), styles["BodyText"])]

    duplicates = payload["duplicates"]
    story.append(KeepTogether([
        Paragraph("Duplicate and overlap candidates", styles["Section"]),
        table([["Candidate type", "Count"],
               ["Exact entity names", len(duplicates["exact_entity_name_candidates"])],
               ["Exact policy statements", len(duplicates["exact_policy_statement_candidates"])],
               ["Near entity names", len(duplicates["near_entity_name_candidates"])]],
              [125 * mm, 45 * mm]),
        Paragraph("Permissions are inferred from the selected cached snapshot. OCI remains the authority for runtime authorization.",
                  styles["BodyText"]),
    ]))

    def footer(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setFillColor(colors.white)
        canvas.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
        canvas.setStrokeColor(colors.HexColor("#B8C2CC"))
        canvas.line(15 * mm, 12 * mm, 195 * mm, 12 * mm)
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#5B6573"))
        canvas.drawString(15 * mm, 8 * mm, "OCI IAM Plotter - read-only cached analysis")
        canvas.drawRightString(195 * mm, 8 * mm, f"Page {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    return stream.getvalue()


def write_report(payload: dict[str, Any], output: Path, snapshot: Snapshot | None = None) -> Path:
    """Write JSON, Markdown, PDF, CSV, or Excel according to the output extension."""
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.suffix.lower() == ".json":
        output.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    elif output.suffix.lower() in {".md", ".markdown"}:
        output.write_text(markdown_report(payload), encoding="utf-8")
    elif output.suffix.lower() == ".csv":
        output.write_text(csv_report(payload), encoding="utf-8")
    elif output.suffix.lower() == ".xlsx":
        if snapshot is None:
            raise ValueError("Excel report output requires the source snapshot")
        output.write_bytes(xlsx_report(snapshot, payload))
    elif output.suffix.lower() == ".pdf":
        output.write_bytes(pdf_report(payload))
    else:
        raise ValueError("Report output must end in .json, .md, .pdf, .csv, or .xlsx")
    return output
