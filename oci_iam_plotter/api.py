"""HTTP-only hosted UI for OCI IAM Plotter.

OCI Generative AI Hosted Applications route normal HTTP requests reliably but
do not proxy Streamlit's relative WebSocket endpoint.  This module deliberately
uses only JSON HTTP endpoints and browser polling.
"""
from __future__ import annotations

from dataclasses import asdict
from io import BytesIO
import json
from pathlib import Path
from typing import Literal
from time import time

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from starlette.middleware.sessions import SessionMiddleware

from .collector import OCICollector
from .analysis import find_duplicates, parse_policy_statement, policy_analysis
from .auth import credentials_match
from .drift import snapshot_drift
from .graph import build_multi_focus_graph, graph_data
from .jobs import collection_logs, collection_status, start_collection_job
from .query import run_iam_agent
from .reporting import csv_report, markdown_report, pdf_report, report_payload, xlsx_report
from .summarizer import OCIReasoner
from .store import SnapshotStore
from .settings import Settings
from .object_store import ObjectSnapshotArchive
from .models import Snapshot

SETTINGS = Settings.from_env()
OBJECT_ARCHIVE = ObjectSnapshotArchive.from_settings(SETTINGS)
STORE = SnapshotStore(SETTINGS.cache_dir, object_archive=OBJECT_ARCHIVE)
STATIC_DIR = Path(__file__).with_name("static")
ACTIVE_PORTAL_SESSIONS: dict[str, float] = {}
SESSION_TTL_SECONDS = 8 * 60 * 60


class CollectionRequest(BaseModel):
    profile_name: str = Field(default="DEFAULT", min_length=1, max_length=100)
    config_text: str = Field(min_length=1)
    pem_text: str = Field(min_length=1)
    use_security_token: bool = False
    security_token_text: str = ""


class LoginRequest(BaseModel):
    username: str
    password: str


class InvestigationRequest(BaseModel):
    focus_ids: list[str] = Field(min_length=1, max_length=12)
    depth: int = Field(default=2, ge=1, le=2)
    relations: list[str] | None = None
    max_edges: int = Field(default=24, ge=4, le=80)


class ReportRequest(BaseModel):
    user_ids: list[str] = Field(default_factory=list, max_length=50)
    format: Literal["json", "markdown", "csv", "xlsx", "pdf"] = "json"
    include_summary: bool = True


class QuestionRequest(BaseModel):
    question: str = Field(min_length=2, max_length=1000)
    history: list[dict[str, str]] = Field(default_factory=list, max_length=12)


class MapExportRequest(BaseModel):
    nodes: list[dict] = Field(min_length=1, max_length=120)
    edges: list[dict] = Field(default_factory=list, max_length=300)
    format: Literal["png", "pdf"] = "png"


class TenancySelection(BaseModel):
    tenancy_id: str = Field(min_length=1)


class BucketSnapshotSelection(BaseModel):
    object_name: str = Field(min_length=8, max_length=1024)


app = FastAPI(title="OCI IAM Plotter", docs_url=None, redoc_url=None)
app.add_middleware(SessionMiddleware, secret_key=__import__("secrets").token_urlsafe(32), https_only=False)
app.mount("/assets", StaticFiles(directory=STATIC_DIR / "assets"), name="assets")


def require_login(request: Request) -> None:
    if not request.session.get("authenticated"):
        raise HTTPException(401, "Sign in first")
    touch_portal_session(request)


def touch_portal_session(request: Request) -> int:
    """Track active browser sessions in this app process without storing identities."""
    now = time()
    expired = [key for key, last_seen in ACTIVE_PORTAL_SESSIONS.items() if now - last_seen > SESSION_TTL_SECONDS]
    for key in expired:
        ACTIVE_PORTAL_SESSIONS.pop(key, None)
    session_id = request.session.get("portal_session_id")
    if request.session.get("authenticated") and session_id:
        ACTIVE_PORTAL_SESSIONS[session_id] = now
    return len(ACTIVE_PORTAL_SESSIONS)


def selected_snapshot_record(request: Request):
    """Return the newest retained snapshot for the tenancy selected in this session."""
    records = STORE.list_history()
    if not records:
        raise HTTPException(404, "No snapshot has been collected")
    tenancy_id = request.session.get("active_tenancy_id")
    matching = [record for record in records if record.tenancy_id == tenancy_id]
    if not matching:
        matching = [records[0]]
        request.session["active_tenancy_id"] = matching[0].tenancy_id
    selected_path = request.session.get("active_snapshot_path")
    if selected_path:
        selected = next((item for item in matching if str(item.path) == selected_path), None)
        if selected:
            return selected
        request.session.pop("active_snapshot_path", None)
    return matching[0]


def loaded_snapshot(request: Request):
    return STORE.load(selected_snapshot_record(request).path)


def selected_tenancy_records(request: Request):
    record = selected_snapshot_record(request)
    return [item for item in STORE.list_history() if item.tenancy_id == record.tenancy_id]


def excel_rows(rows: list[dict], filename: str, sheet_name: str = "Evidence") -> Response:
    """Return a compact Excel workbook for a filtered evidence view."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]
    headers = list(dict.fromkeys(key for row in rows for key in row)) or ["No records"]
    sheet.append(headers)
    for row in rows:
        sheet.append([json.dumps(row.get(key), default=str) if isinstance(row.get(key), (dict, list)) else row.get(key)
                      for key in headers])
    fill = PatternFill("solid", fgColor="17365D")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(48, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    output = BytesIO()
    workbook.save(output)
    return Response(content=output.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def map_export_bytes(nodes: list[dict], edges: list[dict], output_format: str) -> bytes:
    """Render a portable map artifact without requiring browser-only drawing APIs."""
    from PIL import Image, ImageDraw, ImageFont
    from reportlab.lib.colors import HexColor
    from reportlab.pdfgen import canvas

    width, height = 1800, max(900, ((len(nodes) + 4) // 5) * 170 + 180)
    positions = {item.get("id"): (170 + (index % 5) * 365, 120 + (index // 5) * 165)
                 for index, item in enumerate(nodes)}
    colors = {"user": "#2F80ED", "domain_user": "#2F80ED", "group": "#21A179",
              "domain_group": "#21A179", "domain": "#B35AA7", "dynamic_group": "#8B5CF6",
              "policy": "#EF8B3C", "confidential_app": "#E75B76", "oauth_app": "#E75B76",
              "compartment": "#1B9AAA"}
    if output_format == "png":
        image = Image.new("RGB", (width, height), "#07111D")
        draw = ImageDraw.Draw(image)
        font = ImageFont.load_default()
        for edge in edges:
            source, target = positions.get(edge.get("source")), positions.get(edge.get("target"))
            if source and target:
                draw.line([source, target], fill="#83A6BD", width=3)
        for item in nodes:
            x, y = positions[item.get("id")]
            draw.rounded_rectangle((x - 130, y - 32, x + 130, y + 32), radius=13,
                                   fill=colors.get(item.get("kind"), "#526273"), outline="#D8EFFF", width=2)
            label = str(item.get("label", item.get("id", "")))[:28]
            draw.text((x, y - 8), label, fill="white", font=font, anchor="mm")
            draw.text((x, y + 12), str(item.get("kind", "")).replace("_", " "), fill="#D6EDF8", font=font, anchor="mm")
        output = BytesIO(); image.save(output, format="PNG", optimize=True); return output.getvalue()
    output = BytesIO(); pdf = canvas.Canvas(output, pagesize=(width * .4, height * .4)); pdf.setFillColor(HexColor("#07111D")); pdf.rect(0, 0, width * .4, height * .4, fill=1, stroke=0)
    for edge in edges:
        source, target = positions.get(edge.get("source")), positions.get(edge.get("target"))
        if source and target:
            pdf.setStrokeColor(HexColor("#83A6BD")); pdf.line(source[0] * .4, height * .4 - source[1] * .4, target[0] * .4, height * .4 - target[1] * .4)
    for item in nodes:
        x, y = positions[item.get("id")]; color = HexColor(colors.get(item.get("kind"), "#526273")); pdf.setFillColor(color); pdf.roundRect((x - 130) * .4, height * .4 - (y + 32) * .4, 260 * .4, 64 * .4, 10, fill=1, stroke=0); pdf.setFillColor(HexColor("#FFFFFF")); pdf.setFont("Helvetica", 7); pdf.drawCentredString(x * .4, height * .4 - y * .4, str(item.get("label", item.get("id", "")))[:30])
    pdf.showPage(); pdf.save(); return output.getvalue()


@app.post("/api/login")
def login(request: Request, credentials: LoginRequest) -> dict:
    if not credentials_match(credentials.username, credentials.password):
        raise HTTPException(401, "Invalid username or password")
    request.session["authenticated"] = True
    request.session["portal_session_id"] = __import__("secrets").token_urlsafe(18)
    return {"authenticated": True, "active_portal_users": touch_portal_session(request)}


@app.post("/api/logout")
def logout(request: Request) -> dict:
    ACTIVE_PORTAL_SESSIONS.pop(request.session.get("portal_session_id", ""), None)
    request.session.clear()
    return {"authenticated": False, "active_portal_users": len(ACTIVE_PORTAL_SESSIONS)}


@app.get("/api/me")
def me(request: Request) -> dict:
    return {"authenticated": bool(request.session.get("authenticated")), "active_portal_users": touch_portal_session(request)}


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/status")
def status(request: Request) -> dict:
    require_login(request)
    return collection_status()


@app.get("/api/collection-logs")
def logs(request: Request) -> dict:
    require_login(request)
    return {"status": collection_status()["status"], "events": collection_logs()}


@app.post("/api/collect", status_code=202)
def collect(request: Request, body: CollectionRequest) -> dict:
    require_login(request)
    if collection_status()["status"] in {"queued", "running"}:
        raise HTTPException(409, "A collection is already running")
    start_collection_job(
        SETTINGS.cache_dir,
        lambda: OCICollector.from_ephemeral_profile(
            body.config_text, body.pem_text, body.profile_name.strip() or "DEFAULT",
            security_token_text=body.security_token_text, use_security_token=body.use_security_token,
        ), store=STORE,
    )
    return {"status": "queued"}


@app.get("/api/tenancies")
def tenancies(request: Request) -> dict:
    """List tenancies with their independently retained snapshots (newest first)."""
    require_login(request)
    records = STORE.list_history()
    if not records:
        return {"active_tenancy_id": None, "tenancies": []}
    active = selected_snapshot_record(request).tenancy_id
    grouped: dict[str, list] = {}
    for record in records:
        grouped.setdefault(record.tenancy_id, []).append(record)
    items = []
    for tenancy_id, snapshots in grouped.items():
        latest = snapshots[0]
        name = tenancy_id
        try:
            value = STORE.load(latest.path)
            name = next((entity.name for entity in value.entities
                         if entity.id == tenancy_id and entity.name), tenancy_id)
        except (OSError, ValueError, KeyError):
            pass
        items.append({"id": tenancy_id, "name": name,
                      "latest_collected_at": latest.collected_at,
                      "snapshot_count": len(snapshots)})
    return {"active_tenancy_id": active, "tenancies": items}


@app.post("/api/tenancies/select")
def select_tenancy(request: Request, body: TenancySelection) -> dict:
    require_login(request)
    if body.tenancy_id not in {record.tenancy_id for record in STORE.list_history()}:
        raise HTTPException(404, "No retained snapshot exists for this tenancy")
    request.session["active_tenancy_id"] = body.tenancy_id
    request.session.pop("active_snapshot_path", None)
    return {"active_tenancy_id": body.tenancy_id}


def bucket_record_payload(record) -> dict:
    return {"object_name": record.object_name, "tenancy_id": record.tenancy_id,
            "tenancy_name": record.tenancy_name, "collected_at": record.collected_at,
            "source_hash": record.source_hash}


@app.get("/api/bucket-snapshots")
def bucket_snapshots(request: Request) -> dict:
    """List durable archive entries; this does not download snapshot bodies."""
    require_login(request)
    if not OBJECT_ARCHIVE:
        return {"enabled": False, "records": [], "error": "Object Storage archive is disabled."}
    try:
        records = OBJECT_ARCHIVE.list()
        return {"enabled": True, "bucket": OBJECT_ARCHIVE.bucket_name,
                "records": [bucket_record_payload(record) for record in records]}
    except Exception as exc:
        return {"enabled": True, "bucket": OBJECT_ARCHIVE.bucket_name, "records": [],
                "error": f"Unable to list Object Storage collections: {exc}"}


@app.post("/api/bucket-snapshots/select")
def select_bucket_snapshot(request: Request, body: BucketSnapshotSelection) -> dict:
    """Download one explicitly chosen archive entry into the five-item local cache."""
    require_login(request)
    if not OBJECT_ARCHIVE:
        raise HTTPException(503, "Object Storage archive is disabled.")
    if not body.object_name.startswith("tenancies/") or not body.object_name.endswith(".json"):
        raise HTTPException(400, "Invalid bucket snapshot selection.")
    try:
        snapshot = Snapshot.from_dict(OBJECT_ARCHIVE.load(body.object_name))
        path = STORE.save(snapshot, upload_to_object_storage=False)
    except (OSError, ValueError, KeyError, json.JSONDecodeError) as exc:
        raise HTTPException(400, f"Unable to load the selected Object Storage collection: {exc}") from exc
    except Exception as exc:
        raise HTTPException(502, f"Unable to load the selected Object Storage collection: {exc}") from exc
    # Archive file is timestamp/hash-addressed; find it instead of the mutable latest file.
    selected = next((record for record in STORE.list_history()
                     if record.tenancy_id == snapshot.tenancy_id
                     and record.collected_at == snapshot.collected_at
                     and record.source_hash == snapshot.source_hash), None)
    request.session["active_tenancy_id"] = snapshot.tenancy_id
    request.session["active_snapshot_path"] = str(selected.path if selected else path)
    return {"active_tenancy_id": snapshot.tenancy_id, "collected_at": snapshot.collected_at,
            "source_hash": snapshot.source_hash}


@app.get("/api/snapshot")
def snapshot(request: Request) -> dict:
    require_login(request)
    value = loaded_snapshot(request)
    counts: dict[str, int] = {}
    for entity in value.entities:
        counts[entity.kind] = counts.get(entity.kind, 0) + 1
    return {
        "tenancy_id": value.tenancy_id,
        "collected_at": value.collected_at,
        "source_hash": value.source_hash,
        "counts": counts,
        "memberships": len(value.memberships),
        "relationships": len(value.relationships),
        "statements": len(value.statements),
        "stored_users": sum(1 for entity in value.entities if entity.kind == "user"),
        "active_portal_users": touch_portal_session(request),
        "active_users": sum(1 for entity in value.entities if entity.kind == "user"
                            and str(entity.lifecycle_state or "").upper() == "ACTIVE"),
        "warnings": value.warnings,
    }


@app.get("/api/entities")
def entities(request: Request, limit: int = 250) -> list[dict]:
    require_login(request)
    value = loaded_snapshot(request)
    limit = max(1, min(limit, 1000))
    return [asdict(item) for item in value.entities[:limit]]


@app.get("/api/snapshot/download")
def download_snapshot(request: Request) -> FileResponse:
    require_login(request)
    record = selected_snapshot_record(request)
    return FileResponse(record.path, media_type="application/json", filename="oci-iam-snapshot.json")


@app.get("/api/inventory")
def inventory(request: Request, search: str = "", kind: str = "") -> dict:
    require_login(request)
    snapshot = loaded_snapshot(request)
    query = search.casefold().strip()
    rows = []
    for entity in snapshot.entities:
        if kind and entity.kind != kind:
            continue
        if query and query not in json.dumps(asdict(entity), default=str).casefold():
            continue
        rows.append(asdict(entity))
    return {"entities": rows[:1000], "total": len(rows),
            "kinds": sorted({entity.kind for entity in snapshot.entities})}


@app.get("/api/inventory/export")
def export_inventory(request: Request, search: str = "", kind: str = "") -> Response:
    require_login(request)
    snapshot = loaded_snapshot(request)
    query = search.casefold().strip()
    rows = [asdict(entity) for entity in snapshot.entities
            if (not kind or entity.kind == kind)
            and (not query or query in json.dumps(asdict(entity), default=str).casefold())]
    return excel_rows(rows, "oci-iam-filtered-inventory.xlsx", "Filtered inventory")


@app.post("/api/investigate")
def investigate(request: Request, body: InvestigationRequest) -> dict:
    require_login(request)
    snapshot = loaded_snapshot(request)
    try:
        graph = build_multi_focus_graph(
            snapshot, body.focus_ids, depth=body.depth,
            max_nodes=body.max_edges + len(body.focus_ids),
            relations=set(body.relations) if body.relations else None,
            max_edges=body.max_edges,
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    data = graph_data(graph)
    entities = {entity.id: asdict(entity) for entity in snapshot.entities}
    return {**data, "focus_ids": body.focus_ids,
            "available_relations": sorted({edge.get("relation", "RELATED_TO") for edge in data["edges"]}),
            "details": [entities[item] for item in body.focus_ids if item in entities],
            "limitations": ["Solid relationships are collected or parsed evidence; rule-derived possibilities are marked inferred.",
                            "The focused map is capped for readability and is not a tenancy-wide authorization graph."]}


@app.get("/api/user-analysis/{user_id}")
def user_analysis(request: Request, user_id: str) -> dict:
    require_login(request)
    try:
        return policy_analysis(loaded_snapshot(request), user_id)
    except ValueError as exc:
        raise HTTPException(404, str(exc)) from exc


@app.get("/api/policy-statements")
def policy_statements(request: Request, search: str = "") -> dict:
    require_login(request)
    snapshot = loaded_snapshot(request)
    policies = {item.id: item.name for item in snapshot.entities if item.kind == "policy"}
    query = search.casefold().strip()
    rows = []
    for statement in snapshot.statements:
        row = asdict(parse_policy_statement(statement)) | {"policy": policies.get(statement.policy_id, statement.policy_id),
                                                            "policy_id": statement.policy_id}
        if not query or query in json.dumps(row, default=str).casefold():
            rows.append(row)
    return {"statements": rows, "parsed": sum(row["confidence"] == "parsed" for row in rows)}


@app.get("/api/policy-statements/export")
def export_policy_statements(request: Request, search: str = "") -> Response:
    require_login(request)
    snapshot = loaded_snapshot(request)
    policies = {item.id: item.name for item in snapshot.entities if item.kind == "policy"}
    query = search.casefold().strip()
    rows = [asdict(parse_policy_statement(statement)) | {"policy": policies.get(statement.policy_id, statement.policy_id),
                                                          "policy_id": statement.policy_id}
            for statement in snapshot.statements]
    rows = [row for row in rows if not query or query in json.dumps(row, default=str).casefold()]
    return excel_rows(rows, "oci-iam-filtered-policy-statements.xlsx", "Policy statements")


@app.get("/api/duplicates")
def duplicates(request: Request) -> dict:
    require_login(request)
    return find_duplicates(loaded_snapshot(request))


@app.get("/api/duplicates/export")
def export_duplicates(request: Request) -> Response:
    require_login(request)
    result = find_duplicates(loaded_snapshot(request))
    rows = [{"candidate_type": "near_name", "kind": item["kind"], "left": item["left"],
             "right": item["right"], "similarity": item["similarity"]}
            for item in result["near_entity_name_candidates"]]
    rows += [{"candidate_type": "exact_entity_name", **item}
             for group in result["exact_entity_name_candidates"] for item in group]
    rows += [{"candidate_type": "exact_policy_statement", **item}
             for group in result["exact_policy_statement_candidates"] for item in group]
    return excel_rows(rows, "oci-iam-duplicate-candidates.xlsx", "Duplicate candidates")


@app.get("/api/history")
def history(request: Request) -> dict:
    require_login(request)
    records = selected_tenancy_records(request)
    return {"records": [{"index": index, "tenancy_id": item.tenancy_id, "collected_at": item.collected_at,
                          "source_hash": item.source_hash}
                         for index, item in enumerate(records)]}


@app.get("/api/drift")
def drift(request: Request, baseline: int, current: int) -> dict:
    require_login(request)
    records = selected_tenancy_records(request)
    if baseline < 0 or current < 0 or baseline >= len(records) or current >= len(records) or baseline == current:
        raise HTTPException(400, "Choose two different retained collections.")
    try:
        return snapshot_drift(STORE.load(records[baseline].path), STORE.load(records[current].path))
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/report")
def report(request: Request, body: ReportRequest) -> Response:
    require_login(request)
    snapshot = loaded_snapshot(request)
    try:
        analyses = [policy_analysis(snapshot, user_id) for user_id in body.user_ids]
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    summaries = [OCIReasoner().summarize(item) for item in analyses] if body.include_summary and analyses else []
    payload = report_payload(snapshot, analyses, find_duplicates(snapshot), summaries)
    formats = {
        "json": (json.dumps(payload, indent=2), "application/json", "oci-iam-report.json"),
        "markdown": (markdown_report(payload), "text/markdown", "oci-iam-report.md"),
        "csv": (csv_report(payload), "text/csv", "oci-iam-report.csv"),
    }
    if body.format in formats:
        content, media_type, filename = formats[body.format]
        return Response(content=content, media_type=media_type, headers={"Content-Disposition": f'attachment; filename="{filename}"'})
    content = xlsx_report(snapshot, payload) if body.format == "xlsx" else pdf_report(payload)
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if body.format == "xlsx" else "application/pdf"
    filename = "oci-iam-report.xlsx" if body.format == "xlsx" else "oci-iam-report.pdf"
    return Response(content=content, media_type=media_type, headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/ask")
def ask(request: Request, body: QuestionRequest) -> dict:
    require_login(request)
    evidence = run_iam_agent(loaded_snapshot(request), body.question, body.history)
    result = OCIReasoner().answer_question(body.question, evidence)
    return {"answer": result["summary"], "evidence": evidence, "source": result["source"],
            "warning": result.get("warning"), "model_id": result.get("model_id"), "agent": evidence["agent"]}


@app.post("/api/map-export")
def map_export(request: Request, body: MapExportRequest) -> Response:
    require_login(request)
    content = map_export_bytes(body.nodes, body.edges, body.format)
    media_type = "image/png" if body.format == "png" else "application/pdf"
    filename = f"oci-iam-focused-map.{body.format}"
    return Response(content=content, media_type=media_type,
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/")
def index() -> FileResponse:
    return FileResponse(STATIC_DIR / "index.html")
